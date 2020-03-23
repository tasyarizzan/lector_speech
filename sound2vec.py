from utils import speech_to_text, get_volume_times, otnosh, get_monotone
from audio_classification import is_speech
from pydub import AudioSegment
from openpyxl import Workbook


THRESHOLD_LOW = 100
THRESHOLD_LOUD = 100000
THRESHOLD_MONOTONE = 1000
THRESHOLD_SILENSE = 10


def none_to_empty(x):
    if x is None:
        return ''
    else:
        return x


class Sound2VecModel:
    def __init__(self, audio_path):
        sound = AudioSegment.from_wav(audio_path)
        sound = sound.set_channels(1)
        one_channel_path = audio_path[:-4] + '_one.wav'
        sound.export(one_channel_path, format="wav")

        # Setup Excel
        self.__wb = Workbook()
        ws1 = self.__wb.active
        ws1.title = 'Lector Speech Analysis'
        ws1.cell(column=1, row=1, value='Lector Speech Analysis')
        ws1.cell(column=1,row=2, value='Filename:')
        ws1.cell(column=2, row=2, value=audio_path)
        self.is_speech = is_speech(one_channel_path)
        ws1.cell(column=1, row=3, value="Is speech?")
        ws1.cell(column=2, row=3, value=self.is_speech)
        try:
            if self.is_speech:
                print('* ЗАПИСЬ ЯВЛЯЕТСЯ ЗВУКОМ')
                print(self.get_volume(one_channel_path))
                self.volume = self.get_volume(one_channel_path)
                #ws1.cell(column=1, row=4, value="Volume:")
                #ws1.cell(column=2, row=4, value=self.volume)
                print('Громкость: ', self.volume)
                self.silens_segm = self.get_silens_segm(one_channel_path)
                ws1.cell(column=1, row=5, value="Silence segments:")
                ws1.cell(column=2, row=5, value=self.silens_segm)
                print('Сегменты: ', self.silens_segm)
                self.monotone = self.get_monotone(one_channel_path)
                ws1.cell(column=1, row=6, value="Monotone level:")
                ws1.cell(column=2, row=6, value=self.monotone)
                print('Монотонность: ', self.monotone)
                self.text = self.get_text(one_channel_path)
                #ws1.cell(column=1, row=7, value="Text of speech:")
                #ws1.cell(column=2, row=7, value=self.text)
                print('ТЕКСТ: ', self.text)
                self.result = 'Получен вектор с параметрами: '
            else:
                self.result = 'Аудио файл не сожержит речь!'
        except Exception as e:
            self.result = e
            self.volume = None
            self.silens_segm = None
            self.monotone = None
            self.text = None
        excel_path = audio_path[:-4] + "_analysis.xlsx"
        self.__wb.save(excel_path)

    def __str__(self):
        if self.is_speech:
            return f'{self.result} {none_to_empty(self.volume)} {none_to_empty(self.silens_segm)} {none_to_empty(self.monotone)} {none_to_empty(self.text)}'
        else:
            return f'{self.result}'

    def get_volume(self, audio_path):
        print('OL')
        print(get_volume_times(audio_path))
        return get_volume_times(audio_path)

    def get_silens_segm(self, audio_path):
        return otnosh(audio_path, THRESHOLD_SILENSE=10)

    def get_monotone(self, audio_path):
        return get_monotone(audio_path, THRESHOLD_MONOTONE)

    def get_text(self, audio_path):

        return speech_to_text(audio_path)

    def get_json(self):
        pass

    def get_recomend(self):
        pass


test_path = 'audio/test1.wav'
if __name__ == "__main__":
    sound = Sound2VecModel('test1.wav')
    # print(sound)
    # sound.get_text()
